# pylint: disable=invalid-name, anomalous-backslash-in-string

"""
Script to save Daily (or Weekly, Monthly, etc.) reports in their proper locations in the host computer.
The code below could work for any daily report routines.
You can make any changes so that it can work for e.g. monthly reports.

Usage:

- Download and copy all excel files to a folder in the host machine.
- Open a bash shell and navigate at the directory where the python script is located
- Run the script through:
  `python merge_reports.py`

"""

import re
import logging
import datetime
from typing import List
from copy import copy
from os.path import isfile, join, isdir
from os import listdir, remove, getcwd
from shutil import copy2
from path import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def merge_daily_report_parts(reports: List[str], folder: str):
    """
    This function takes a list of excel names for daily reports
    that are to be merged according to the order they are given.
    (The second is appended to the first, etc.)

    It produces one final excel file that has the name specified and removes
    the other files!

    Args:
        reports: List of strings of the filenames of the excel files that
            are to be merged.
        folder: Folder where the excel files are located.
    """
    # create list of workbooks/worksheets
    workbooks = []
    worksheets = []
    for idx, val in enumerate(reports):
        fpath = join(folder, val)
        logging.info("Loading sheet %s.", val)
        workbooks.append(load_workbook(fpath))
        worksheets.append(workbooks[idx]["hardcoded_name_of_the_sheet"])
        logging.info("Removing sheet %s.", val)
        remove(fpath)

    for idx, sheet in enumerate(worksheets):
        # python indexing starts from 0. We start from the first sheet
        # hence we don't need to append/change anything here.
        if idx == 0:
            continue

        for row in sheet.iter_rows(min_row=3):
            new_row = []
            for cell in row:
                tmp = Cell(worksheets[0])       # Get an empty cell
                tmp.value = copy(cell.value)
                tmp._style = copy(cell._style)
                new_row.append(tmp)
            worksheets[0].append(new_row)

    # save file
    report_name = get_entry(reports[0]) + ".xlsx"
    logging.info("Saving merged daily excel report at:\n%s", report_name)
    workbooks[0].save(join(folder, report_name))
    workbooks[0].close()


def get_entry(name: str) -> str:
    """
    This function splits the excel filename in order to isolate the part
    that defines the entity that the excel file has data for. This allows
    identification of all excel files that describe data for the same report.
    In this example, the identification is done through the date (e.g the 
    excel filename should have the following format: excel_filename_20220122.xlsx)
    
    This function is useful when downloading batches of excel files with the same name.

    Example using re function:
    >>> a = 'asdadada12345678b'
    >>> m = re.search(r"[0-9]{8}", a)
    >>> m.end(0)
    16
    >>> a[:16]
    'asdadada12345678'

    Args:
        name: string of the excel filename

    Returns:
        The entry that corresponds to the given filename.
    """
    search_obj = re.search(r"[0-9]{8}", name)
    if search_obj:
        # get index of last match (x.end(0) and split string!)
        return name[: search_obj.end(0)]
    else:
        return ""


def find_daily_parts(excel_list: List[str]) -> List[List[str]]:
    """
    This function is given a list of excel paths. It will output a list
    where each element is an ordered list of items that need merging
    to form a complete daily (or weekly, monthly, etc.) report.

    Args:
        excel_list: List of fimenames to be identified as partial daily reports.

    Returns:
        The list of lists that correspond to partial daily reports on the relevant
        excel files.
    """

    unique_entries = set()
    duplicate_entries = set()

    for el in excel_list:
        entry = get_entry(el)
        if entry in unique_entries:
            duplicate_entries.add(entry)
        elif entry == "":
            logging.error("File %s resulted in missing report entity!", el)
            raise ValueError("Unable to parse report entity!")
        else:
            unique_entries.add(entry)

    
    master_list = []
    for dup in duplicate_entries:
        logging.info("Identified partial daily report %s", dup)
        temp_list = []
        for el in excel_list:
            if get_entry(el) == dup:
                temp_list.append(el)
        master_list.append(temp_list)

    return master_list


def order_excel_list(listc: List[str], folder: str) -> List[str]:
    """
    Give the list of daily partial excel reports in descending chronological order.
    ***For this to work, there should exist a column in the excel files which contains dates.

    Args:
        listc: list of partial daily reports.
        folder: Folder where the excel files are located.
    """

    reference = {}
    dates = set()

    for it in listc:

        fpath = join(folder, it)
        workbook = load_workbook(fpath)
        
        worksheet = workbook["hardcoded_name_of_the_sheet"]
        # Hardcoded cell given that we know the expected format of the excel workbooks
        cell = worksheet["A2"]      # In this example, column "A" contains the dates and they start from cell "A2".
        date1 = datetime.datetime.strptime(cell.value, "%d-%b-%Y")

        reference[date1] = it
        # In case it happens that you accidentally download 
        # duplicate data from the relevant server
        if date1 in dates:
            logging.error(
                "Duplicate date found in excel sheet! Re-create files:\n %s", listc
            )
            raise ValueError(
                f"Duplicate date found in excel sheet! Re-create files:\n{listc}"
            )
        dates.add(date1)

    date_list = list(dates)
    # Sort the dates in whatever chronological order you prefer
    date_list.sort(reverse=True)
     
    # Get the partial excel files in a chronological
    # order in order to merge them correctly
    return_list = []
    for q in date_list:
        return_list.append(reference[q])

    return return_list


def merge_daily_partial_reports(
    master_list: List[List[str]], folder: str
):
    """
    This function accepts the master_list that contains the list of partial rad reports.
    It orders each list appropriately and performs the collation.
    """

    for ilist in master_list:
        new_list = order_excel_list(ilist, folder)
        logging.info("Starting collation of partial daily reports.")

        merge_daily_report_parts(new_list, folder)


def get_excel_files_in_folder(folder_dir: Path) -> List[str]:
    """
    Gets all files within a directory that are .xlsx files

    Args:
        folder_dir: folder path where the reports are located.

    Retruns:
        List of filepaths for excel reports.

    """
    # check if folder exists:
    if not isdir(folder_dir):
        logging.error("%s directory not found", folder_dir)
        raise NotADirectoryError(f"{folder_dir} directory not found")

    # get list of entries
    possible_entries = listdir(folder_dir)

    reports_list = []
    for el in possible_entries:
        if isfile(join(folder_dir, el)):
            if el.endswith(".xlsx"):
                reports_list.append(el)

    return reports_list


def get_project_identifier_from_name(name: str) -> str:
    """
    The function reads the excel filename of the expected format
    and gives back the project identifier.
    This function can be modified depending on the format of the excel filenames.

    For example:
    PROJECT_NAME_20220117.xlsx returns PROJECT_NAME_
    etc
    """
    parts = name.split("_")
    return "_".join(parts[:2])


def get_destination_folder(project_identifier: str) -> str:
    """
    This function gets the project identifier of a specific excel file
    and computes the destination folder.
    Note that the location is HARDCODED!
    So the directories need to have a specific name format.
    """
    # ensure project_id is lowercase
    project_identifier = project_identifier.lower()

    return f"C:\\Projects\\{project_identifier}\\directory_name\\you_would_like\\to_put_the_file\\"


def rename_excel(name: str) -> str:
    """
    Rename the excel files if necessary
    """
    return name.replace("_something_", "_somethingElse_")


# Move existing files to destination folders
def merge_and_move_reports(folder_dir: str):
    """
    Main function to handle tasks on excel files
    """

    # Get list of files
    reports_list = get_excel_files_in_folder(folder_dir)
    logging.info("Got reports list.")

    partials_list = find_daily_parts(reports_list)
    merge_daily_partial_reports(partials_list, folder_dir)
    logging.info("Partial reports merged!")

    # Get new report list after collation
    reports_list = get_excel_files_in_folder(folder_dir)

    logging.info(
        "Using hardcoded location [D:\\Projects\\$project\\dir1\\dir2\\dir3\\] for destination!"
    )
    # Move reports
    for rprt in reports_list:
        # get destination folder for report:
        prj_id = get_project_identifier_from_name(rprt)
        destination = get_destination_folder(prj_id)
        # rename if needed
        new_name = rename_excel(rprt)
        fl_from = join(folder_dir, rprt)
        fl_to = join(destination, new_name)
        logging.info("Copying file %s to %s", fl_from, fl_to)
        copy2(fl_from, fl_to)


if __name__ == "__main__":

    logging.basicConfig(
        level=logging.DEBUG,
        # Write logs both to file and stdout!
        handlers=[
            logging.FileHandler("daily_reports.log", mode="w"), # file
            logging.StreamHandler(),                        # sys.stdout and sys.stderr
        ],
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    logging.info("Starting Script to handle Excel Daily reports.")

    # Hardcoded folder
    reports_folder = Path(r"C:\Users\the_user\Reports")
    # Run the main function
    merge_and_move_reports(reports_folder)
